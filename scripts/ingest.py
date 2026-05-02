#!/usr/bin/env python3
"""
ITBI SP — Bronze ingestion script
Loads monthly XLSX sheets into DuckDB bronze_itbi_transactions.
Contract version: v1.0 (2026-05-01)

Privacy guarantee: columns 'Cartório de Registro' and 'Matrícula do Imóvel'
are excluded by name before any row is written to DuckDB.
"""

import re
import sys
import logging
from datetime import datetime, timezone
from pathlib import Path

import duckdb
import openpyxl
from dotenv import load_dotenv
import os

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_PATTERN = re.compile(r'^[A-Z]{3}-\d{4}$')

RETAINED_COLUMNS = [
    "sql_cadastro",
    "nome_logradouro",
    "numero",
    "complemento",
    "bairro",
    "referencia",
    "cep",
    "natureza_transacao",
    "valor_transacao",
    "data_transacao",
    "valor_venal_referencia",
    "proporcao_transmitida_pct",
    "valor_venal_proporcional",
    "base_calculo",
    "tipo_financiamento",
    "valor_financiado",
    "situacao_sql",
    "area_terreno_m2",
    "testada_m",
    "fracao_ideal",
    "area_construida_m2",
    "uso_iptu_codigo",
    "uso_iptu_descricao",
    "padrao_iptu_codigo",
    "padrao_iptu_descricao",
    "acc_iptu",
]

# Map from source Excel header names to internal field names
SOURCE_TO_FIELD = {
    "N° do Cadastro (SQL)":                      "sql_cadastro",
    "Nome do Logradouro":                         "nome_logradouro",
    "Número":                                     "numero",
    "Complemento":                                "complemento",
    "Bairro":                                     "bairro",
    "Referência":                                 "referencia",
    "CEP":                                        "cep",
    "Natureza de Transação":                      "natureza_transacao",
    "Valor de Transação (declarado pelo contribuinte)": "valor_transacao",
    "Data de Transação":                          "data_transacao",
    "Valor Venal de Referência":                  "valor_venal_referencia",
    "Proporção Transmitida (%)":                  "proporcao_transmitida_pct",
    "Valor Venal de Referência (proporcional)":   "valor_venal_proporcional",
    "Base de Cálculo adotada":                    "base_calculo",
    "Tipo de Financiamento":                      "tipo_financiamento",
    "Valor Financiado":                           "valor_financiado",
    "Situação do SQL":                            "situacao_sql",
    "Área do Terreno (m2)":                       "area_terreno_m2",
    "Testada (m)":                                "testada_m",
    "Fração Ideal":                               "fracao_ideal",
    "Área Construída (m2)":                       "area_construida_m2",
    "Uso (IPTU)":                                 "uso_iptu_codigo",
    "Descrição do uso (IPTU)":                    "uso_iptu_descricao",
    "Padrão (IPTU)":                              "padrao_iptu_codigo",
    "Descrição do padrão (IPTU)":                "padrao_iptu_descricao",
    "ACC (IPTU)":                                 "acc_iptu",
    # EXCLUDED — never read:
    # "Cartório de Registro"   → excluded
    # "Matrícula do Imóvel"    → excluded
}

PRIVACY_COLUMNS = {"Cartório de Registro", "Matrícula do Imóvel"}

BRONZE_TABLE = "raw_itbi_transactions"

CREATE_BRONZE_DDL = f"""
CREATE TABLE IF NOT EXISTS {BRONZE_TABLE} (
    sql_cadastro              VARCHAR,
    nome_logradouro           VARCHAR,
    numero                    VARCHAR,
    complemento               VARCHAR,
    bairro                    VARCHAR,
    referencia                VARCHAR,
    cep                       VARCHAR,
    natureza_transacao        VARCHAR,
    valor_transacao           VARCHAR,
    data_transacao            VARCHAR,
    valor_venal_referencia    VARCHAR,
    proporcao_transmitida_pct VARCHAR,
    valor_venal_proporcional  VARCHAR,
    base_calculo              VARCHAR,
    tipo_financiamento        VARCHAR,
    valor_financiado          VARCHAR,
    situacao_sql              VARCHAR,
    area_terreno_m2           VARCHAR,
    testada_m                 VARCHAR,
    fracao_ideal              VARCHAR,
    area_construida_m2        VARCHAR,
    uso_iptu_codigo           VARCHAR,
    uso_iptu_descricao        VARCHAR,
    padrao_iptu_codigo        VARCHAR,
    padrao_iptu_descricao     VARCHAR,
    acc_iptu                  VARCHAR,
    source_sheet              VARCHAR NOT NULL,
    ingested_at               TIMESTAMP NOT NULL
);
"""

INSERT_SQL = f"""
INSERT INTO {BRONZE_TABLE} (
    sql_cadastro, nome_logradouro, numero, complemento, bairro,
    referencia, cep, natureza_transacao, valor_transacao, data_transacao,
    valor_venal_referencia, proporcao_transmitida_pct, valor_venal_proporcional,
    base_calculo, tipo_financiamento, valor_financiado, situacao_sql,
    area_terreno_m2, testada_m, fracao_ideal, area_construida_m2,
    uso_iptu_codigo, uso_iptu_descricao, padrao_iptu_codigo, padrao_iptu_descricao,
    acc_iptu, source_sheet, ingested_at
) VALUES (
    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
    ?, ?, ?, ?, ?, ?, ?, ?
)
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def setup_logging() -> logging.Logger:
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log_file = log_dir / f"ingestion_{ts}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_file),
        ],
    )
    logger = logging.getLogger("ingest")
    logger.info(f"Log file: {log_file}")
    return logger


def clean_value(val):
    """Strip whitespace; convert empty/None to NULL."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Core ingestion logic
# ---------------------------------------------------------------------------

def ingest_sheet(
    ws,
    sheet_name: str,
    con: duckdb.DuckDBPyConnection,
    ingested_at: datetime,
    logger: logging.Logger,
) -> int:
    """Ingest one worksheet. Returns row count inserted."""

    # Read header row (row 1)
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    # Build column index map: field_name → col index (0-based)
    col_index: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header in PRIVACY_COLUMNS:
            logger.info(
                f"  [PRIVACY] Column '{header}' detected in sheet {sheet_name} — SKIPPED (never read)"
            )
            continue
        field = SOURCE_TO_FIELD.get(header)
        if field:
            col_index[field] = idx

    # Verify all required retained columns are present
    missing = [f for f in RETAINED_COLUMNS if f not in col_index]
    if missing:
        logger.warning(
            f"  [WARN] Sheet {sheet_name}: missing source columns: {missing}. "
            "They will be NULL in Bronze."
        )

    # Delete existing rows for this sheet (idempotency)
    result = con.execute(
        f"SELECT COUNT(*) FROM {BRONZE_TABLE} WHERE source_sheet = ?", [sheet_name]
    ).fetchone()
    deleted_count = result[0] if result else 0
    if deleted_count > 0:
        con.execute(f"DELETE FROM {BRONZE_TABLE} WHERE source_sheet = ?", [sheet_name])
        logger.info(f"  Deleted {deleted_count} existing rows for sheet '{sheet_name}' (re-ingestion)")

    # Read data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        record = []
        for field in RETAINED_COLUMNS:
            idx = col_index.get(field)
            val = clean_value(row[idx]) if idx is not None else None
            record.append(val)

        record.append(sheet_name)           # source_sheet
        record.append(ingested_at)          # ingested_at

        rows.append(tuple(record))

    if rows:
        con.executemany(INSERT_SQL, rows)
    logger.info(f"  Inserted {len(rows)} rows for sheet '{sheet_name}'")
    return len(rows)


def run_ingestion(xlsx_path: str, db_path: str, logger: logging.Logger) -> None:
    logger.info(f"Opening workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    matching_sheets = [s for s in wb.sheetnames if SHEET_PATTERN.match(s)]
    if not matching_sheets:
        logger.error(f"No sheets matching {SHEET_PATTERN.pattern} found in workbook. Aborting.")
        sys.exit(1)
    logger.info(f"Sheets to ingest: {matching_sheets}")

    con = duckdb.connect(db_path)
    con.execute(CREATE_BRONZE_DDL)

    ingested_at = datetime.now(timezone.utc)
    total_rows = 0

    for sheet_name in matching_sheets:
        logger.info(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        count = ingest_sheet(ws, sheet_name, con, ingested_at, logger)
        total_rows += count

    con.close()
    wb.close()
    logger.info(f"Ingestion complete. Total rows inserted: {total_rows}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    load_dotenv()

    logger = setup_logging()

    parser = argparse.ArgumentParser(description="ITBI SP Bronze ingestion")
    parser.add_argument(
        "--file",
        default="data/GUIAS DE ITBI PAGAS (2).xlsx",
        help="Path to the ITBI XLSX workbook",
    )
    args = parser.parse_args()

    db_path = os.environ.get("DB_PATH")
    if not db_path:
        logger.error("DB_PATH environment variable is not set. Create a .env file.")
        sys.exit(1)

    xlsx_path = args.file
    if not Path(xlsx_path).exists():
        logger.error(f"XLSX file not found: {xlsx_path}")
        sys.exit(1)

    run_ingestion(xlsx_path, db_path, logger)
